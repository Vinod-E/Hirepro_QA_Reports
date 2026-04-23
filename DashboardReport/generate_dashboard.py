import os
import re
import datetime
import json
import html
import xlrd
import subprocess
import re
from pathlib import Path
from openpyxl import load_workbook
import pandas as pd

from Config import configfile
from DashboardReport.utils import sanitize_content

# Config from central source
REPORTS_DIR = Path(configfile.REPORT_DIR)
OUTPUT_FILE = Path(configfile.DASHBOARD_REPORT)
MASTER_LIST_FILE = Path(configfile.MASTER_REPORTS_LIST)
TARGET_EXECUTION_GOAL = configfile.TARGET_EXECUTION_GOAL
EXPECTED_REPORT_COUNT = configfile.EXPECTED_REPORT_COUNT


def get_report_folders(base_dir, limit=5):
    """Find the last 'limit' date folders (YYYYMMDD) in reverse chronological order."""
    if not base_dir.exists(): return []
    folders = []
    for d in base_dir.iterdir():
        if d.is_dir() and re.match(r'^\d{8}$', d.name):
            folders.append(d.name)
    folders.sort(reverse=True)
    return folders[:limit]


def create_excel_preview(filepath):
    """Generate a clean HTML preview of an Excel file."""
    try:
        html_path = filepath.parent / (filepath.name + ".html")
        # Read without assuming header to show all rows (including metadata/off-row headers)
        df = pd.read_excel(filepath, header=None)
        # Fill NA values with 'NA' string
        df = df.fillna('NA')
        table_html = df.to_html(classes='preview-table', index=False, header=False)

        table_html = sanitize_content(table_html)

        filename = filepath.name
        html_content = f"""<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>Preview: {filename}</title>
    <link href="https://fonts.googleapis.com/css2?family=Outfit:wght@400;600;700;800&display=swap" rel="stylesheet">
    <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">
    <style>
        :root {{ --primary: #4f46e5; --bg: #f8fafc; --text: #1e293b; --border: #e2e8f0; --success: #10b981; --danger: #ef4444; }}
        body {{ font-family: 'Outfit', sans-serif; background-color: var(--bg); color: var(--text); padding: 2rem; margin: 0; }}
        .container {{ max-width: 1400px; margin: 0 auto; background: white; padding: 2.5rem; border-radius: 1.5rem; box-shadow: 0 10px 25px -5px rgba(0,0,0,0.05); border: 1px solid var(--border); }}
        header {{ display: flex; justify-content: space-between; align-items: center; margin-bottom: 2rem; padding-bottom: 1.5rem; border-bottom: 2px solid var(--bg); }}
        .title-area {{ display: flex; align-items: center; gap: 1rem; }}
        .title-area i {{ font-size: 2rem; color: var(--primary); }}
        h1 {{ font-size: 1.5rem; font-weight: 800; margin: 0; }}
        .actions {{ display: flex; gap: 1rem; }}
        .btn {{ text-decoration: none; padding: 0.6rem 1.25rem; border-radius: 10px; font-weight: 700; font-size: 0.85rem; display: flex; align-items: center; gap: 0.5rem; transition: all 0.2s; border: none; cursor: pointer; }}
        .btn-download {{ background: var(--primary); color: white; }}
        .btn-download:hover {{ transform: translateY(-2px); box-shadow: 0 4px 12px rgba(79, 70, 229, 0.3); }}
        .btn-back {{ background: #f1f5f9; color: var(--text); }}
        .btn-back:hover {{ background: #e2e8f0; }}
        .table-wrapper {{ overflow-x: auto; border-radius: 12px; border: 1px solid var(--border); background: white; }}
        .preview-table {{ width: 100%; border-collapse: collapse; font-size: 0.85rem; border-spacing: 0; }}
        .preview-table td {{ padding: 0.85rem 1.15rem; border-bottom: 1px solid var(--border); border-right: 1px solid var(--border); white-space: nowrap; color: var(--text); }}
        .preview-table tr:hover {{ background: #f1f5f9 !important; }}

        /* Header & Sub-header Styling */
        .preview-table tr:first-child td,
        .preview-table tr.sub-header td {{ 
            background: #f1f5f9 !important; 
            font-weight: 800 !important; 
            color: #475569 !important; 
            text-transform: uppercase; 
            font-size: 0.75rem; 
            letter-spacing: 0.025em;
            border-bottom: 2px solid #cbd5e1 !important;
        }}

        .preview-table tr:nth-child(even) {{ background: #fafbfc; }}
        .preview-table tr:hover {{ background: #f1f5f9 !important; }}

        /* Status Colors */
        .status-pass {{ color: var(--success) !important; font-weight: 800; background: #ecfdf5 !important; }}
        .status-fail {{ color: var(--danger) !important; font-weight: 800; background: #fef2f2 !important; }}
        .status-na {{ color: #94a3b8; font-style: italic; opacity: 0.5; }}
    </style>
</head>
<body>
    <div class="container">
        <header>
            <div class="title-area">
                <i class="fas fa-file-excel"></i>
                <div>
                    <h1>{filename}</h1>
                    <div style="font-size: 0.75rem; color: #64748b; font-weight: 600; margin-top: 0.2rem; text-transform: uppercase;">Spreadsheet Preview (All Rows)</div>
                </div>
            </div>
            <div class="actions">
                <button onclick="window.close()" class="btn btn-back"><i class="fas fa-times"></i> Close Tab</button>
                <a href="{filename}" download="{filename}" class="btn btn-download"><i class="fas fa-download"></i> Download Excel</a>
            </div>
        </header>
        <div class="table-wrapper">
            {table_html}
        </div>
    </div>
    <script>
        document.querySelectorAll('.preview-table tr').forEach((tr, rowIndex) => {{
            let statusCount = 0;
            const cells = tr.querySelectorAll('td');

            cells.forEach(td => {{
                const text = td.innerText.trim().toUpperCase();

                // Highlight Pass/Fail
                if (['PASS', 'SUCCESS', 'PASSED'].includes(text)) {{
                    td.classList.add('status-pass');
                }} else if (['FAIL', 'FAILURE', 'FAILED', 'ERROR'].includes(text)) {{
                    td.classList.add('status-fail');
                }} else if (text === 'NA') {{
                    td.classList.add('status-na');
                }}

                // Header detection keywords
                if (['STATUS', 'IDENTITY', 'ID'].includes(text)) {{
                    statusCount++;
                }}
            }});

            // If row has multiple 'Status' labels, it's a sub-header
            if (statusCount >= 2 && rowIndex > 0) {{
                tr.classList.add('sub-header');
            }}
        }});
    </script>
</body>
</html>"""
        html_path.write_text(html_content, encoding='utf-8')
        return f"{filepath.name}.html"
    except Exception as e:
        print(f"Error creating preview for {filepath}: {e}")
        return None


def scan_folder(folder_path, date_str):
    """Scan a specific folder for report files and extract metrics."""
    reports = []
    for f in folder_path.iterdir():
        if not f.is_file(): continue
        if f.suffix not in ('.html', '.xls', '.xlsx'): continue
        if f.name.endswith('.xls.html') or f.name.endswith('.xlsx.html'): continue

        mod_ts = f.stat().st_mtime
        entry = {
            "name": f.name,
            "path": f"reports/{date_str}/{f.name}",
            "type": "HTML" if f.suffix == '.html' else "Excel",
            "mod_time": datetime.datetime.fromtimestamp(mod_ts).strftime('%Y-%m-%d %H:%M:%S'),
            "mod_time_ts": mod_ts,
            "summary": {}
        }
        if f.suffix == '.html':
            with f.open('r', encoding='utf-8', errors='ignore') as file:
                original_content = file.read()
                entry["summary"] = extract_counts(original_content[:1000000], f.name)

            # Sanitize original HTML report in-place
            sanitized_content = sanitize_content(original_content)
            if sanitized_content != original_content:
                with f.open('w', encoding='utf-8') as file:
                    file.write(sanitized_content)

            entry["view_path"] = entry["path"]
        else:
            entry["summary"] = extract_excel_counts(f)
            preview_file = create_excel_preview(f)
            entry["view_path"] = f"reports/{date_str}/{preview_file}" if preview_file else entry["path"]
        reports.append(entry)
    return reports


# Pre-compiled regex patterns for speed
NEWMAN_REQUESTS_RE = re.compile(r'Total Requests <span class="badge badge-light">(\d+)</span>')
NEWMAN_FAILED_RE = re.compile(r'Failed Tests <span class="badge badge-light">(\d+)</span>')
NEWMAN_COLLECTION_RE = re.compile(r'<strong> Collection:</strong>\s*(.*?)\s*<br>')
MOCHAWESOME_RE = re.compile(r'data-raw="(.*?)"')
TITLE_RE = re.compile(r'<title>(.*?)</title>')
PYTEST_HTML_RE = re.compile(r'id="data-container"\s+data-jsonblob="(.*?)"')


def extract_counts(html_content, filename):
    """Extract test metrics from HTML reports (Newman, Mochawesome, or Pytest)."""
    data = {"requests": 0, "failed": 0, "skipped": 0, "collection": "Unknown", "pass_percent": 0, "failed_cases": []}
    clean_html = lambda text: html.unescape(text).strip() if text else ""

    if "pytest-html" in html_content:
        m_pytest = PYTEST_HTML_RE.search(html_content)
        if m_pytest:
            try:
                json_str = html.unescape(m_pytest.group(1))
                report_data = json.loads(json_str)
                tests = report_data.get("tests", {})

                total = 0
                failed = 0
                failed_cases = []
                for test_id, results in tests.items():
                    # pytest-html stores results in a list (usually one per test)
                    for result in results:
                        total += 1
                        if result.get("result") in ["Failed", "Error"]:
                            failed += 1
                            failed_cases.append(test_id.split('::')[-1])

                data["requests"] = total
                data["failed"] = failed
                data["failed_cases"] = failed_cases
                data["collection"] = "Pytest Suite"
                if total > 0:
                    data["pass_percent"] = round(((total - failed) / total) * 100, 1)
                return data
            except:
                pass

    if "Newman Run Dashboard" in html_content or "newman-report" in html_content:
        m_req = NEWMAN_REQUESTS_RE.search(html_content);
        m_fail = NEWMAN_FAILED_RE.search(html_content);
        m_coll = NEWMAN_COLLECTION_RE.search(html_content)
        if m_req: data["requests"] = int(m_req.group(1))
        if m_fail: data["failed"] = int(m_fail.group(1))
        if m_coll: data["collection"] = sanitize_content(clean_html(m_coll.group(1)))
        if data["requests"] > 0: data["pass_percent"] = round(
            ((data["requests"] - data["failed"]) / data["requests"]) * 100, 1)
        return data
    if "mochawesome" in html_content:
        m_raw = MOCHAWESOME_RE.search(html_content)
        if m_raw:
            try:
                stats = json.loads(html.unescape(m_raw.group(1))).get('stats', {})
                data["requests"], data["failed"], data["pass_percent"] = stats.get('tests', 0), stats.get('failures',
                                                                                                          0), stats.get(
                    'passPercent', 0)
                m_title = TITLE_RE.search(html_content)
                if m_title: data["collection"] = sanitize_content(clean_html(m_title.group(1)))
                return data
            except:
                pass
    return data


def extract_excel_counts(filepath):
    """Extract test metrics from Excel reports (.xls or .xlsx)."""
    data = {"requests": 0, "failed": 0, "collection": "Excel Data Asset", "pass_percent": 0}
    try:
        suffix = filepath.suffix.lower()
        if suffix == '.xls':
            workbook = xlrd.open_workbook(filepath)
            sheet = workbook.sheet_by_index(0)
            all_rows = [sheet.row_values(i) for i in range(sheet.nrows)]
        else:
            wb = load_workbook(filepath, data_only=True, read_only=True)
            sheet = wb.active
            all_rows = [[cell.value for cell in row] for row in sheet.iter_rows()]
            wb.close()

        if not all_rows: return data
        row_values = all_rows[0]
        p_val, f_val = 0, 0
        pass_headers = {"no.of test cases", "success cases", "passed cases", "total pass steps", "sessions pass",
                        "total pass"}
        fail_headers = {"failed test cases", "failure cases", "failed cases", "total fail steps", "failed count",
                        "total fail"}

        found_summary = False
        for i, val in enumerate(row_values):
            s = str(val).strip().lower() if val is not None else ""
            if any(h in s for h in pass_headers):
                if i + 1 < len(row_values) and isinstance(row_values[i + 1], (int, float)):
                    p_val = int(row_values[i + 1])
                    found_summary = True
            elif any(h in s for h in fail_headers):
                if i + 1 < len(row_values) and isinstance(row_values[i + 1], (int, float)):
                    f_val = int(row_values[i + 1])
                    found_summary = True

        if not found_summary:
            status_idx = -1
            for i, val in enumerate(row_values):
                if str(val).strip().lower() == 'status':
                    status_idx = i
                    break

            if status_idx != -1:
                for row in all_rows[1:]:
                    if status_idx < len(row):
                        status_val = str(row[status_idx]).strip().upper()
                        if status_val == 'PASS':
                            p_val += 1
                        elif status_val == 'FAIL':
                            f_val += 1

        total = p_val + f_val
        data["requests"], data["failed"] = total, f_val
        if total > 0: data["pass_percent"] = round(((total - f_val) / total) * 100, 1)
        if row_values and isinstance(row_values[0], str): data["collection"] = sanitize_content(row_values[0])
    except Exception as e:
        print(f"Error reading Excel {filepath}: {e}")
    return data


def get_git_commit():
    """Retrieve current Git commit hash."""
    try:
        return subprocess.check_output(['git', 'rev-parse', '--short', 'HEAD'], stderr=subprocess.DEVNULL).decode(
            'ascii').strip()
    except:
        return "Unknown"


def classify_report(filename):
    """Determine the group category for a report based on its filename."""
    upper_name = filename.upper()
    if upper_name.endswith('_PW.XLSX'): return "AI - Playwright Reports"
    if upper_name.endswith('_CLAUDE.XLSX'): return "AI - Claude Reports"
    if upper_name.startswith('API_'): return "API Reports"
    if upper_name.startswith('UI_SSO_'): return "SSO Reports"
    if upper_name.startswith('UI_') and 'SLOT' in upper_name: return "SLOTS Reports"
    if upper_name.endswith('_MS.XLS') or upper_name.endswith('_MS.XLSX'): return "Microsite Reports"
    if upper_name.startswith('UI_'): return "CRPO Reports"
    if upper_name.startswith('SPRINT_'): return "ATS Reports"
    return "Cypress Reports"


def generate_styles():
    """Return the CSS style block."""
    return """
    <style>
        :root {
            --primary: #4f46e5; --primary-dark: #4338ca; --bg: #f8fafc; --card-bg: #ffffff;
            --text: #1e293b; --text-muted: #64748b; --success: #10b981; --danger: #ef4444;
            --border: #e2e8f0; --shadow: 0 10px 15px -3px rgb(0 0 0 / 0.1), 0 4px 6px -4px rgb(0 0 0 / 0.1);
            --glass: rgba(255, 255, 255, 0.7);
        }
        * { margin: 0; padding: 0; box-sizing: border-box; font-family: 'Outfit', sans-serif; -webkit-tap-highlight-color: transparent; }
        html, body { -ms-overflow-style: none; scrollbar-width: none; overflow-x: hidden; scroll-behavior: smooth; }
        html::-webkit-scrollbar, body::-webkit-scrollbar { display: none; }
        body { background-color: var(--bg); color: var(--text); padding: 1.5rem 1rem; width: 100%; min-height: 100vh; touch-action: manipulation; }
        .container { max-width: 1240px; margin: 0 auto; }
        header { margin-bottom: 1.5rem; display: grid; grid-template-columns: 1fr auto 1fr; align-items: center; padding-bottom: 1rem; border-bottom: 1px solid var(--border); }
        h1 { font-size: 1.4rem; font-weight: 800; color: var(--text); line-height: 1.2; }
        .subtitle { color: var(--text-muted); font-size: 0.75rem; margin-top: 0.25rem; font-weight: 600; text-transform: uppercase; letter-spacing: 0.05em; }

        .header-home-btn { display: none; text-decoration: none; align-items: center; justify-content: center; width: 44px; height: 44px; border-radius: 12px; background: #f1f5f9; color: var(--primary); transition: all 0.2s; margin-bottom: 0.5rem; }
        .header-home-btn:active { background: #e2e8f0; transform: scale(0.95); }

        .hero-banner { 
            background: var(--card-bg); border: 1px solid var(--border); padding: 2.5rem; border-radius: 2rem; 
            margin-bottom: 2.5rem; box-shadow: var(--shadow); display: flex; align-items: center; gap: 3rem; 
            position: relative; overflow: hidden; transform: translateZ(0);
        }
        .hero-banner.stale-banner { border-color: #fca5a5; background: #fffafb; box-shadow: 0 10px 25px -5px rgba(239, 68, 68, 0.08); }
        .hero-banner.stale-banner::before { background: linear-gradient(to bottom, #ef4444, #fca5a5); }
        .hero-main { flex: 1; min-width: 0; }
        .hero-label { font-size: 0.85rem; font-weight: 700; text-transform: uppercase; color: var(--text-muted); margin-bottom: 0.5rem; letter-spacing: 0.05em; display: flex; align-items: center; gap: 8px; flex-wrap: wrap; line-height: 1.6; }
        .hero-value { font-size: 3rem; font-weight: 900; color: var(--text); display: flex; align-items: baseline; gap: 0.5rem; }
        .sync-badge { 
            background: #ffe4e6; color: #e11d48; padding: 4px 12px; border-radius: 99px; 
            font-size: 0.65rem; font-weight: 800; border: 1px solid #fecdd3; 
            display: inline-flex; align-items: center; gap: 6px; 
            animation: softPulse 2s infinite ease-in-out; 
            white-space: nowrap; margin-left: 4px;
        }
        @keyframes softPulse { 0%, 100% { opacity: 1; transform: scale(1); } 50% { opacity: 0.8; transform: scale(0.97); } }
        .hero-target { font-size: 1.25rem; color: var(--text-muted); font-weight: 500; }
        .progress-container { flex: 1.5; }
        .progress-header { display: flex; justify-content: space-between; margin-bottom: 0.75rem; font-weight: 700; font-size: 0.9rem; }
        .progress-outer { height: 16px; background: #f1f5f9; border-radius: 8px; overflow: hidden; }
        .progress-inner { height: 100%; background: linear-gradient(90deg, #4f46e5, #818cf8); transition: width 1s cubic-bezier(0.4, 0, 0.2, 1); }

        .stats-grid { display: grid; grid-template-columns: repeat(auto-fit, minmax(240px, 1fr)); gap: 1.5rem; margin-bottom: 3rem; }
        .stat-card { background: linear-gradient(145deg, #ffffff, #f9fafb); border: 1px solid var(--border); padding: 1.75rem; border-radius: 1.5rem; display: flex; align-items: center; gap: 1.5rem; transition: all 0.4s cubic-bezier(0.175, 0.885, 0.32, 1.275); box-shadow: 0 10px 20px -10px rgba(0,0,0,0.1); position: relative; overflow: hidden; }
        .stat-card:hover { transform: translateY(-8px) scale(1.02); box-shadow: 0 20px 30px -15px rgba(0,0,0,0.15); border-color: rgba(79, 70, 229, 0.4); background: #ffffff; }
        .stat-card::before { content: ''; position: absolute; top: 0; left: 0; width: 4px; height: 100%; opacity: 0.6; }
        .stat-card.blue::before { background: #4f46e5; } .stat-card.green::before { background: #10b981; }
        .stat-card.red::before { background: #ef4444; } .stat-card.orange::before { background: #f59e0b; }
        .stat-icon { width: 52px; height: 52px; border-radius: 1rem; display: flex; align-items: center; justify-content: center; font-size: 1.5rem; box-shadow: 0 4px 6px -1px rgba(0,0,0,0.05); }
        .icon-blue { background: #eff6ff; color: #3b82f6; } .icon-green { background: #ecfdf5; color: #10b981; }
        .icon-red { background: #fff5f5; color: #ef4444; } .icon-orange { background: #fffbeb; color: #f59e0b; }
        .stat-info .value { font-size: 2rem; font-weight: 900; display: block; color: var(--text); line-height: 1.1; }
        .stat-info .label { color: var(--text-muted); font-size: 0.75rem; font-weight: 700; text-transform: uppercase; letter-spacing: 0.05em; margin-top: 0.25rem; }

        .alert-row { margin-bottom: 2rem; }
        .alert-card { background: #fff5f5; border: 1px solid #fed7d7; padding: 1rem 1.5rem; border-radius: 1rem; display: flex; align-items: center; justify-content: space-between; color: #c53030; }
        .alert-card i { font-size: 1.25rem; margin-right: 0.75rem; }

        details { background: var(--card-bg); border: 1px solid var(--border); border-radius: 16px; margin-bottom: 1.5rem; box-shadow: var(--shadow); position: relative; overflow: visible; transition: transform 0.2s cubic-bezier(0.4, 0, 0.2, 1); }
        details:hover { border-color: #cbd5e1; }
        summary { padding: 1.5rem; font-weight: 700; font-size: 1.1rem; cursor: pointer; background: #f8fafc; display: flex; justify-content: space-between; align-items: center; list-style: none; position: relative; border-radius: 15px 15px 0 0; transition: background 0.2s; -webkit-tap-highlight-color: transparent; }
        details:not([open]) summary { border-radius: 15px; }
        summary::after { content: '\\f078'; font-family: "Font Awesome 6 Free"; font-weight: 900; transition: transform 0.3s cubic-bezier(0.175, 0.885, 0.32, 1.275); margin-left: 1rem; color: var(--text-muted); font-size: 0.9rem; }
        details[open] summary::after { transform: rotate(180deg); color: var(--primary); }
        details[open] summary { border-bottom: 1px solid var(--border); margin-bottom: 0; }

        .content-wrapper { display: grid; grid-template-rows: 0fr; transition: grid-template-rows 0.4s cubic-bezier(0.4, 0, 0.2, 1); }
        details[open] .content-wrapper { grid-template-rows: 1fr; }
        .content-inner { overflow: hidden; }

        summary:hover { background: #f1f5f9 !important; }
        summary span { transition: transform 0.2s; }
        summary:hover span { transform: translateX(5px); }

        table { width: 100%; border-collapse: collapse; }
        th { background: #f8fafc; padding: 1rem 1.5rem; color: var(--text-muted); font-weight: 700; font-size: 0.8rem; text-transform: uppercase; text-align: left; border-bottom: 1px solid var(--border); }
        td { padding: 1.25rem 1.5rem; border-bottom: 1px solid var(--border); vertical-align: middle; }
        .report-link { color: var(--text); text-decoration: none; font-weight: 600; display: flex; align-items: center; gap: 0.75rem; }
        .badge { padding: 0.5rem 0.75rem; border-radius: 8px; font-size: 0.7rem; font-weight: 700; }
        .badge-html { background: #e0e7ff; color: #4338ca; } .badge-excel { background: #ecfdf5; color: #065f46; }
        .chip { font-size: 0.8rem; padding: 6px 16px; border-radius: 20px; font-weight: 900; background: #f1f5f9; white-space: nowrap; display: inline-flex; align-items: center; justify-content: center; min-width: 70px; box-shadow: 0 2px 4px rgba(0,0,0,0.05); }
        .chip.success { background: #d1fae5; color: #065f46; } .chip.danger { background: #fee2e2; color: #991b1b; }
        .chip i { margin-right: 0.35rem; font-size: 0.7rem; opacity: 0.7; }

        .search-area { margin-bottom: 2rem; position: relative; }
        .search-area input { 
            width: 100%; padding: 1rem 3.5rem; border-radius: 14px; border: 1.5px solid var(--border); 
            outline: none; transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1); background: white;
            font-size: 0.95rem; font-weight: 500;
        }
        .search-area input:focus { border-color: var(--primary); box-shadow: 0 0 0 4px rgba(79, 70, 229, 0.1); }
        .search-area i { position: absolute; left: 1.4rem; top: 50%; transform: translateY(-50%); color: var(--text-muted); z-index: 10; pointer-events: none; transition: color 0.3s; }
        .search-area input:focus + i { color: var(--primary); }

        footer { margin-top: 4rem; padding: 2rem 0; border-top: 1px solid var(--border); display: flex; justify-content: space-between; align-items: center; color: var(--text-muted); font-size: 0.85rem; font-weight: 500; }
        .commit-badge { background: #f1f5f9; padding: 0.25rem 0.6rem; border-radius: 6px; font-family: monospace; font-weight: 700; color: var(--text); border: 1px solid var(--border); font-size: 0.75rem; }

        /* Liquid Glass Navigation */
        .main-nav { 
            display: flex; position: relative; justify-content: center; gap: 0; 
            margin-bottom: 2.5rem; background: rgba(241, 245, 249, 0.75); 
            backdrop-filter: blur(12px); -webkit-backdrop-filter: blur(12px);
            padding: 0.45rem; border-radius: 99px; width: fit-content; margin: 0 auto 2.5rem auto;
            border: 1px solid rgba(255,255,255,0.6); box-shadow: 0 8px 32px rgba(0,0,0,0.06);
        }
        .nav-btn { 
            position: relative; z-index: 2; padding: 0.75rem 2.8rem; border-radius: 99px; 
            border: none; background: transparent; color: var(--text-muted); 
            font-weight: 700; cursor: pointer; transition: color 0.3s, transform 0.1s; 
            display: flex; align-items: center; gap: 0.75rem; font-size: 0.95rem; 
            touch-action: manipulation;
        }
        .nav-btn:active { transform: scale(0.96); }
        .nav-btn.active { color: var(--primary); }
        .nav-btn i { font-size: 1rem; transition: transform 0.3s; }
        .nav-btn.active i { transform: scale(1.1); }
        .nav-indicator {
            position: absolute; height: calc(100% - 0.9rem); top: 0.45rem; left: 0.45rem;
            background: white; border-radius: 99px; z-index: 1;
            transition: all 0.35s cubic-bezier(0.23, 1, 0.32, 1);
            box-shadow: 0 4px 15px rgba(79, 70, 229, 0.18);
            width: 0;
            pointer-events: none;
        }

        .tab-content { display: none; }
        .tab-content.active { display: block; }

        /* Date Switching Styles */
        .day-view { display: none; transform: translateY(10px); opacity: 0; transition: all 0.3s ease; }
        .day-view.active { display: block; transform: translateY(0); opacity: 1; }
        .view-btn { 
            padding: 0.45rem 1.15rem; border-radius: 99px; 
            border: 1px solid rgba(79, 70, 229, 0.15); 
            background: rgba(255, 255, 255, 0.6); 
            backdrop-filter: blur(10px); -webkit-backdrop-filter: blur(10px);
            color: var(--primary); font-weight: 700; cursor: pointer; 
            transition: all 0.2s cubic-bezier(0.4, 0, 0.2, 1); 
            font-size: 0.75rem; position: relative; overflow: hidden;
            box-shadow: 0 2px 8px rgba(0,0,0,0.03);
            white-space: nowrap;
            touch-action: manipulation;
        }
        .view-btn:active { transform: translateY(0) scale(0.95); }
        .view-btn::before {
            content: ''; position: absolute; top: -50%; left: -150%; width: 200%; height: 200%;
            background: radial-gradient(circle, rgba(255,255,255,0.5) 0%, transparent 70%);
            transition: all 0.6s; pointer-events: none;
        }
        .view-btn:hover::before { left: -50%; top: -50%; }
        .view-btn:hover { 
            background: rgba(255, 255, 255, 0.8); 
            border-color: rgba(79, 70, 229, 0.3); 
            transform: translateY(-2px);
            box-shadow: 0 6px 15px rgba(79, 70, 229, 0.1);
        }
        .view-btn.active { 
            background: var(--primary); color: white; 
            box-shadow: 0 8px 25px rgba(79, 70, 229, 0.35); 
            border-color: var(--primary); 
            transform: translateY(0);
        }
        .view-btn.active::before { display: none; }
        .date-selector-wrapper { margin-bottom: 2rem; display: flex; align-items: center; gap: 1rem; }
        .date-select-input { padding: 0.6rem 1rem; border-radius: 12px; border: 1px solid var(--border); font-weight: 600; font-family: 'Outfit'; color: var(--text); outline: none; background: white; cursor: pointer; }
        .health-index { font-size: 0.7rem; font-weight: 800; padding: 2px 8px; border-radius: 4px; margin-left: 8px; }
        .health-good { background: #d1fae5; color: #065f46; } .health-fair { background: #fffbeb; color: #92400e; } .health-bad { background: #fee2e2; color: #991b1b; }

        /* Custom Tooltip Styles - Bottom Placement */
        [data-tooltip] { position: relative; cursor: pointer; }
        [data-tooltip]::before {
            content: attr(data-tooltip); position: absolute; bottom: 120%; left: 50%;
            transform: translateX(-50%) translateY(5px);
            padding: 0.5rem 0.85rem; background: #1e293b; color: #ffffff;
            font-size: 0.7rem; font-weight: 700; border-radius: 12px; white-space: nowrap;
            opacity: 0; visibility: hidden; transition: all 0.3s cubic-bezier(0.4, 0, 0.2, 1);
            pointer-events: none; box-shadow: 0 10px 20px rgba(0,0,0,0.2); z-index: 10000;
            border: 1px solid rgba(255,255,255,0.1); text-transform: uppercase; letter-spacing: 0.05em;
        }
        [data-tooltip]:hover::before { opacity: 1; visibility: visible; transform: translateX(-50%) translateY(0); }

        .floating-reset[data-tooltip]::before, .floating-home[data-tooltip]::before { bottom: 125%; }

        @keyframes attentionBlink { 0% { opacity: 1; transform: scale(1); } 50% { opacity: 0.7; transform: scale(0.95); } 100% { opacity: 1; transform: scale(1); } }
        .badge-urgent { animation: attentionBlink 1.5s infinite ease-in-out; background: white; color: var(--danger); padding: 0.35rem 0.85rem; border-radius: 99px; font-weight: 800; font-size: 0.75rem; text-transform: uppercase; letter-spacing: 0.02em; box-shadow: 0 2px 4px rgba(239, 68, 68, 0.2); }
        .sync-label { font-size: 0.65rem; font-weight: 800; color: #d97706; display: flex; align-items: center; gap: 4px; margin-top: 2px; text-transform: uppercase; }
        .sync-label.error { color: #ea580c; }

        /* History Section */
        .history-section { background: var(--card-bg); border: 1px solid var(--border); border-radius: 1.5rem; padding: 1.5rem; box-shadow: var(--shadow); margin-bottom: 2.5rem; overflow: hidden; }
        .trend-title { font-size: 1rem; font-weight: 800; margin-bottom: 1.25rem; display: flex; align-items: center; gap: 0.75rem; color: var(--text); }
        .trend-title i { color: var(--primary); }

        .history-table-wrapper { width: 100%; overflow-x: auto; -webkit-overflow-scrolling: touch; margin: 0 -1.5rem; padding: 0 1.5rem; width: calc(100% + 3rem); }
        .history-table { width: 100%; border-collapse: collapse; font-size: 0.85rem; min-width: 580px; }
        .history-table th { padding: 0.75rem; text-align: left; color: var(--text-muted); border-bottom: 1px solid var(--border); background: transparent; white-space: nowrap; }
        .history-table td { padding: 0.75rem; border-bottom: 1px solid var(--border); vertical-align: middle; }
        .history-date { font-weight: 700; color: var(--text); white-space: nowrap; }

        .floating-home {
            position: fixed; bottom: 2rem; right: 2rem; width: 56px; height: 56px; 
            background: var(--primary); color: white; border-radius: 20px; 
            display: flex; align-items: center; justify-content: center;
            box-shadow: 0 10px 30px rgba(79, 70, 229, 0.3);
            z-index: 9999; border: 1px solid rgba(255,255,255,0.2);
            transition: all 0.4s cubic-bezier(0.175, 0.885, 0.32, 1.275);
            text-decoration: none; font-size: 1.4rem;
            backdrop-filter: blur(8px);
        }
        .floating-home:hover { transform: scale(1.1) translateY(-5px); box-shadow: 0 15px 40px rgba(79, 70, 229, 0.45); background: var(--primary-dark); }
        .floating-home i { transition: transform 0.3s ease; }
        .floating-home:hover i { transform: rotate(-10deg); }
        .floating-home[data-tooltip]:hover::before { transform: translateX(-50%) translateY(0); }

        .floating-reset {
            position: fixed; bottom: 2rem; right: 6.5rem; width: 56px; height: 56px; 
            background: var(--primary); color: white; border-radius: 20px; 
            display: flex; align-items: center; justify-content: center;
            box-shadow: 0 10px 30px rgba(79, 70, 229, 0.3); cursor: pointer;
            z-index: 9999; border: 1px solid rgba(255,255,255,0.2); transition: all 0.3s cubic-bezier(0.175, 0.885, 0.32, 1.275);
            text-decoration: none; font-size: 1.2rem;
            backdrop-filter: blur(8px);
        }
        .floating-reset:hover { transform: scale(1.1) translateY(-5px); box-shadow: 0 15px 40px rgba(79, 70, 229, 0.45); }
        .floating-reset:hover i { transform: rotate(45deg); }
        .floating-reset i { transition: transform 0.3s; }
        .floating-reset:active { transform: scale(0.9); }
        .floating-reset[data-tooltip]:hover::before { transform: translateX(-50%) translateY(0); }

        @media (max-width: 768px) {
            body { padding: 1rem 0.5rem; overflow-x: hidden; }
            .container { width: 100%; padding: 0; }
            header { grid-template-columns: 1fr; gap: 1rem; }
            header > div { width: 100%; display: flex; flex-direction: column; align-items: center; justify-content: center; text-align: center !important; }
            h1 { font-size: 1.2rem; }
            .subtitle { font-size: 0.65rem; }

            .hero-banner { flex-direction: column; gap: 1.5rem; text-align: center; padding: 1.5rem 1.25rem; margin: 0 auto 1.5rem auto; align-items: center !important; }
            .hero-main { width: 100%; display: flex; flex-direction: column; align-items: center; }
            .hero-label { font-size: 0.75rem; justify-content: center !important; width: 100%; }
            .hero-value { font-size: 2rem; justify-content: center !important; width: 100%; margin: 0.5rem 0; }
            .hero-target { font-size: 1rem; }
            .progress-container { width: 100%; }
            .progress-header { font-size: 0.8rem; justify-content: space-between !important; }
            .hero-banner.stale-banner { text-align: center !important; }

            .alert-card { flex-direction: column; text-align: center; gap: 1rem; padding: 1.25rem 1rem; }
            .alert-card span { font-size: 0.85rem; line-height: 1.4; }
            .badge-urgent { font-size: 0.65rem; padding: 0.3rem 0.6rem; }

            .stats-grid { 
                grid-template-columns: repeat(2, 1fr); 
                background: white; 
                border: 1px solid var(--border); 
                border-radius: 1.25rem; 
                padding: 0.5rem; 
                gap: 0; 
                margin-bottom: 2rem;
            }
            .stat-card { 
                flex-direction: column; 
                text-align: center; 
                padding: 1.25rem 0.25rem; 
                background: transparent !important; 
                border: none !important; 
                box-shadow: none !important; 
            }
            .stat-icon { width: 36px; height: 36px; font-size: 1rem; margin-bottom: 0.4rem; }
            .stat-info .value { font-size: 1.25rem; }
            .stat-info .label { font-size: 0.6rem; }
            .stat-card:nth-child(1), .stat-card:nth-child(2) { border-bottom: 1px solid #f1f5f9 !important; }
            .stat-card:nth-child(1), .stat-card:nth-child(3) { border-right: 1px solid #f1f5f9 !important; }

            .main-nav { padding: 0.35rem; width: 95%; max-width: 400px; margin-bottom: 1.5rem; gap: 0.2rem; }
            .nav-btn { padding: 0.75rem 0.5rem; font-size: 0.8rem; flex: 1; justify-content: center; gap: 0.5rem; }
            .nav-indicator { height: calc(100% - 0.7rem); top: 0.35rem; }

            .history-section { padding: 0.75rem; border-radius: 1rem; overflow: visible; }
            .history-table-wrapper { margin: 0; padding: 0; width: 100%; overflow-x: auto; -webkit-overflow-scrolling: touch; }
            .history-table { font-size: 0.7rem; }

            .search-area input { font-size: 0.9rem; padding: 0.85rem 1rem 0.85rem 3.25rem; }

            .report-group { width: 100%; border-radius: 14px; }
            summary { padding: 1.25rem 3rem 1.25rem 1.25rem; position: relative; }
            summary::after { position: absolute; right: 1.25rem; top: 50%; transform: translateY(-50%); }
            details[open] summary::after { transform: translateY(-50%) rotate(180deg); }
            summary > div { flex-direction: column; align-items: center; text-align: center; gap: 0.75rem !important; width: 100%; }
            summary span { font-size: 0.95rem !important; font-weight: 800; margin-right: 0; display: block; width: 100%; }
            summary div .chip { font-size: 0.75rem; min-width: 60px; padding: 6px 12px; }
            summary div div { justify-content: center !important; width: 100% !important; margin: 0 !important; }

            .content-inner { overflow-x: auto; width: 100%; }
            table { min-width: 550px; }
            td, th { padding: 0.75rem 0.5rem; font-size: 0.75rem; }
            .view-btn { padding: 0.35rem 0.8rem; font-size: 0.65rem; }

            footer { flex-direction: column; gap: 1rem; text-align: center; font-size: 0.75rem; }
            .header-home-btn { display: flex !important; }
            .floating-home, .floating-reset { display: none !important; }
            [data-tooltip]::before { display: none !important; }
        }
        .loader-overlay {
            position: fixed; top: 0; left: 0; right: 0; bottom: 0;
            background: rgba(255,255,255,0.85); backdrop-filter: blur(12px); -webkit-backdrop-filter: blur(12px);
            display: flex; align-items: center; justify-content: center; z-index: 99999;
            transition: opacity 0.3s ease-out;
        }
        .loader-spinner {
            width: 32px; height: 32px; border: 3px solid #e2e8f0;
            border-top: 3px solid var(--primary); border-radius: 50%;
            animation: spin 0.6s linear infinite;
        }
        @keyframes spin { 0% { transform: rotate(0deg); } 100% { transform: rotate(360deg); } }
        @keyframes pulse-text { 0%, 100% { opacity: 1; } 50% { opacity: 0.5; } }
        .loader-text { 
            font-weight: 800; color: var(--primary); font-size: 0.8rem; 
            letter-spacing: 0.1em; text-transform: uppercase; 
            margin-top: 1rem; animation: pulse-text 1.5s ease-in-out infinite; 
        }
        .collapsible-loader {
            display: none; width: 16px; height: 16px; border: 2px solid rgba(0,0,0,0.05); border-top: 2px solid var(--primary); border-radius: 50%; animation: spin 0.6s linear infinite; margin-right: 0.75rem; flex-shrink: 0;
        }
        summary.loading .collapsible-loader { display: inline-block; }
        summary.loading { background: #f1f5f9 !important; }
        @media print {
            .main-nav, .floating-reset, .loader-overlay, .search-area, .view-btn, .nav-btn, header { display: none !important; }
            .container { width: 100%; max-width: 100%; padding: 0; }
            .history-section, .hero-banner { box-shadow: none; border: 1px solid #eee; }
            .day-view { display: block !important; opacity: 1 !important; transform: none !important; }
        }
        .fail-badge { 
            background: #fff1f2; color: #e11d48; padding: 4px 12px; border-radius: 99px; 
            font-size: 0.9rem; font-weight: 800; border: 1px solid #ffe4e6; 
            display: inline-flex; align-items: center; gap: 6px; 
            margin-top: 0.75rem;
            animation: softPulse 2s infinite ease-in-out;
            white-space: nowrap;
        }
    </style>
    """


def generate_landing_page():
    """Generate the dashboard.html landing page for report type selection."""
    index_path = Path(configfile.DASHBOARD_REPORT).parent / "dashboard.html"
    html_content = """<!DOCTYPE html>
<html lang="en">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <script>
        // Immediate theme detection to prevent flash
        (function() {
            const savedTheme = localStorage.getItem('theme');
            const prefersDark = window.matchMedia('(prefers-color-scheme: dark)').matches;
            const isMobile = window.innerWidth < 768;
            
            // If saved theme exists, use it. Otherwise, follow system preference.
            // On mobile, we default to Light if no specific preference is found.
            if (savedTheme === 'dark' || (!savedTheme && prefersDark)) {
                document.documentElement.classList.add('dark');
                document.documentElement.classList.remove('light');
            } else {
                document.documentElement.classList.remove('dark');
                document.documentElement.classList.add('light');
            }
        })();
    </script>
    <link rel="icon" type="image/png" href="https://lh3.googleusercontent.com/proxy/98it6d0vGqaPttxE8ImrHhVvaz5XpPeZ7qiXHcnm9PiPrwTBGvZcWp2vdQVdgZt5b7Vfu7kXnkh6mrKs_q_JIE5GGlw8uRAOeSvJOEtgNXWrXgOoGjGFCnKCA1gITLLZKNxv1mV6szDHEnNLK7RbJnfk-eFMZPlGXRpU2iKeBGr2Gm3-i_Lnv-IVisLlJwRR55nNMx9HndfYnmLOlraDHGgp9Rc7V4pNO1N8S1ZugYR5SUVMi8K_WGFwvYWsEh2cEnW6x-Zw-ncSM27yX509d6pmuUvfohenPAxHMNORCeWO">
    <title>QAInsights | Report Selection</title>
    <script src="https://cdn.tailwindcss.com"></script>
    <script src="https://unpkg.com/lucide@latest"></script>
    <script>
        tailwind.config = {
            darkMode: 'class',
            theme: {
                extend: {
                    colors: {
                        darkBg: '#0f172a',
                        lightBg: '#f8fafc',
                    }
                }
            }
        }
    </script>
    <style>
        @import url('https://fonts.googleapis.com/css2?family=Inter:wght@300;400;600;700&display=swap');
        html, body { 
            scrollbar-width: none;
            -ms-overflow-style: none;
            overflow-x: hidden;
            scroll-behavior: smooth;
        }
        html::-webkit-scrollbar, body::-webkit-scrollbar {
            display: none;
        }
        body { font-family: 'Inter', sans-serif; }

        .card-transition { transition: all 0.4s cubic-bezier(0.175, 0.885, 0.32, 1.275); }

        /* Glassmorphism & Liquid UI */
        .glass-card {
            backdrop-filter: blur(20px) saturate(180%);
            -webkit-backdrop-filter: blur(20px) saturate(180%);
        }

        .bg-lightBg { background-image: radial-gradient(at 0% 0%, rgba(59, 130, 246, 0.03) 0, transparent 50%), radial-gradient(at 50% 0%, rgba(168, 85, 247, 0.03) 0, transparent 50%); }
        .bg-darkBg { background-image: radial-gradient(at 0% 0%, rgba(59, 130, 246, 0.1) 0, transparent 50%), radial-gradient(at 50% 0%, rgba(168, 85, 247, 0.1) 0, transparent 50%); }

        /* Glassmorphism & Gradients */
        .perf-card { 
            background: linear-gradient(135deg, rgba(59, 130, 246, 0.08), rgba(234, 179, 8, 0.08));
            border: 1px solid rgba(234, 179, 8, 0.2);
        }
        .light .perf-card { background: rgba(255, 255, 255, 0.6); border: 1px solid rgba(234, 179, 8, 0.25); }

        .ui-card { 
            background: linear-gradient(135deg, rgba(34, 197, 94, 0.08), rgba(59, 130, 246, 0.08));
            border: 1px solid rgba(34, 197, 94, 0.2);
        }
        .light .ui-card { background: rgba(255, 255, 255, 0.6); border: 1px solid rgba(34, 197, 94, 0.25); }

        .sprint-card { 
            background: linear-gradient(135deg, rgba(99, 102, 241, 0.08), rgba(168, 85, 247, 0.08));
            border: 1px solid rgba(99, 102, 241, 0.2);
        }
        .light .sprint-card { background: rgba(255, 255, 255, 0.6); border: 1px solid rgba(99, 102, 241, 0.25); }

        .liquid-blob {
            position: fixed;
            z-index: -1;
            filter: blur(80px);
            opacity: 0.15;
            border-radius: 50%;
            animation: float 20s infinite alternate ease-in-out;
        }

        @keyframes float {
            0% { transform: translate(0, 0) scale(1); }
            33% { transform: translate(30px, -50px) scale(1.1); }
            66% { transform: translate(-20px, 20px) scale(0.9); }
            100% { transform: translate(0, 0) scale(1); }
        }

        @keyframes spin-slow {
            from { transform: rotate(0deg); }
            to { transform: rotate(360deg); }
        }
        .animate-spin-slow {
            animation: spin-slow 3s linear infinite;
        }
    </style>
</head>
<body class="bg-lightBg dark:bg-darkBg text-slate-900 dark:text-slate-100 min-h-screen transition-colors duration-500 overflow-x-hidden overflow-y-auto">

    <!-- Liquid Background Blobs -->
    <div class="liquid-blob bg-blue-400 w-96 h-96 -top-20 -left-20 opacity-10 dark:opacity-20"></div>
    <div class="liquid-blob bg-purple-400 w-[30rem] h-[30rem] top-1/2 -right-20 opacity-10 dark:opacity-20" style="animation-delay: -5s;"></div>
    <div class="liquid-blob bg-orange-300 w-80 h-80 bottom-0 left-1/4 opacity-10 dark:opacity-20" style="animation-delay: -10s;"></div>

    <div class="fixed top-4 right-4 lg:top-6 lg:right-6 z-50">
        <button onclick="toggleTheme()" class="p-3 rounded-full bg-white dark:bg-slate-800 shadow-xl border border-slate-200 dark:border-slate-700 hover:scale-110 active:scale-95 transition-all">
            <i id="theme-icon" data-lucide="sun" class="w-5 h-5 text-yellow-500"></i>
        </button>
    </div>

    <div class="flex flex-col items-center justify-between min-h-screen p-4 lg:p-6 py-8 lg:py-12 relative z-10">

        <header class="w-full max-w-full lg:max-w-6xl text-center mb-6 lg:mb-12 px-4">
            <div class="flex flex-col items-center justify-center gap-4 lg:gap-6 mb-4 lg:mb-8">
                <img src="https://hirepro.in/wp-content/uploads/2025/05/HirePro-logo.svg" alt="HirePro Logo" class="h-8 lg:h-14 w-auto dark:brightness-200">
                <h1 class="text-2xl lg:text-5xl font-black tracking-tighter text-slate-900 dark:text-white">QA<span class="font-extralight text-slate-500">Insights</span></h1>
            </div>
        </header>


        <main class="w-full max-w-full lg:max-w-5xl space-y-12 lg:space-y-12 px-2 lg:px-0 flex-1">

            <div class="space-y-3 lg:space-y-8">
                <div class="flex items-center gap-4 lg:gap-6">
                    <div class="h-px flex-1 bg-slate-300 dark:bg-slate-700"></div>
                    <h2 class="text-[10px] lg:text-xs font-black uppercase tracking-[0.1em] lg:tracking-[0.4em] text-slate-900 dark:text-slate-100 text-center">Daily Execution Results</h2>
                    <div class="h-px flex-1 bg-slate-300 dark:bg-slate-700"></div>
                </div>

                <div class="grid grid-cols-1 lg:grid-cols-2 gap-3 lg:gap-8">
                    <section onclick="window.location.href='automationreports.html'" 
                             class="ui-card glass-card card-transition rounded-2xl lg:rounded-3xl p-3 lg:p-8 shadow-sm hover:shadow-2xl hover:-translate-y-2 cursor-pointer flex flex-col h-full relative overflow-hidden group">
                        <!-- Card Inner Glow -->
                        <div class="absolute inset-0 bg-gradient-to-br from-white/20 to-transparent opacity-0 group-hover:opacity-100 transition-opacity pointer-events-none"></div>

                        <div class="flex justify-center mb-1 lg:mb-8 relative z-10">
                            <h3 class="text-[11px] lg:text-xs font-bold tracking-widest text-slate-500 dark:text-slate-400 uppercase">Automation Test Report</h3>
                        </div>

                        <div class="grid grid-cols-3 gap-2 lg:gap-4 mb-2 lg:mb-10 flex-1 relative z-10">
                            <div class="flex flex-col items-center text-center group/item">
                                <div class="p-2 lg:p-3 rounded-lg lg:rounded-2xl bg-green-500/10 mb-2 lg:mb-3 group-hover/item:scale-110 transition-transform">
                                    <i data-lucide="check-circle" class="w-5 h-5 lg:w-8 lg:h-8 text-green-600 dark:text-green-500"></i>
                                </div>
                                <span class="text-[9px] lg:text-[11px] font-semibold text-slate-500 dark:text-slate-400 leading-tight">Success Rate</span>
                            </div>
                            <div class="flex flex-col items-center text-center group/item">
                                <div class="p-2 lg:p-3 rounded-lg lg:rounded-2xl bg-orange-500/10 mb-2 lg:mb-3 group-hover/item:scale-110 transition-transform">
                                    <i data-lucide="play-circle" class="w-5 h-5 lg:w-8 lg:h-8 text-orange-500 dark:text-orange-400"></i>
                                </div>
                                <span class="text-[9px] lg:text-[11px] font-semibold text-slate-500 dark:text-slate-400 leading-tight">Total Tests</span>
                            </div>
                            <div class="flex flex-col items-center text-center group/item">
                                <div class="p-2 lg:p-3 rounded-lg lg:rounded-2xl bg-blue-500/10 mb-2 lg:mb-3 group-hover/item:scale-110 transition-transform">
                                    <i data-lucide="monitor" class="w-5 h-5 lg:w-8 lg:h-8 text-blue-600 dark:text-blue-400"></i>
                                </div>
                                <span class="text-[9px] lg:text-[11px] font-semibold text-slate-500 dark:text-slate-400 leading-tight">Browser(Chrome)<br>Coverage</span>
                            </div>
                        </div>

                        <a href="automationreports.html" 
                           onclick="event.stopPropagation();"
                           class="flex items-center justify-center gap-2 w-full bg-green-500/10 dark:bg-green-400/10 text-green-600 dark:text-green-400 font-bold py-3 lg:py-4 rounded-xl lg:rounded-2xl text-[10px] lg:text-xs tracking-widest hover:bg-green-500/20 dark:hover:bg-green-400/20 border border-green-500/20 transition-all uppercase mt-auto relative z-10">
                            View Automation Test Report
                            <i data-lucide="external-link" class="w-4 h-4"></i>
                        </a>
                    </section>

                    <section onclick="window.location.href='performance_daily.html'"
                             class="perf-card glass-card card-transition rounded-2xl lg:rounded-3xl p-3 lg:p-8 shadow-sm hover:shadow-2xl hover:-translate-y-2 cursor-pointer flex flex-col h-full relative overflow-hidden group">
                        <!-- Card Inner Glow -->
                        <div class="absolute inset-0 bg-gradient-to-br from-white/20 to-transparent opacity-0 group-hover:opacity-100 transition-opacity pointer-events-none"></div>

                        <div class="flex justify-center mb-1 lg:mb-8 relative z-10">
                            <h3 class="text-[11px] lg:text-xs font-bold tracking-widest text-slate-500 dark:text-slate-400 uppercase">Performance Report - Daily</h3>
                        </div>

                        <div class="grid grid-cols-3 gap-2 lg:gap-4 mb-2 lg:mb-10 flex-1 relative z-10">
                            <div class="flex flex-col items-center text-center group/item">
                                <div class="p-2 lg:p-3 rounded-lg lg:rounded-2xl bg-yellow-500/10 mb-2 lg:mb-3 group-hover/item:scale-110 transition-transform">
                                    <i data-lucide="timer" class="w-5 h-5 lg:w-8 lg:h-8 text-yellow-600 dark:text-yellow-500"></i>
                                </div>
                                <span class="text-[9px] lg:text-[11px] font-semibold text-slate-500 dark:text-slate-400 leading-tight">Analyze API<br>Performance</span>
                            </div>
                            <div class="flex flex-col items-center text-center group/item">
                                <div class="p-2 lg:p-3 rounded-lg lg:rounded-2xl bg-blue-500/10 mb-2 lg:mb-3 group-hover/item:scale-110 transition-transform">
                                    <i data-lucide="gauge" class="w-5 h-5 lg:w-8 lg:h-8 text-blue-600 dark:text-blue-400"></i>
                                </div>
                                <span class="text-[9px] lg:text-[11px] font-semibold text-slate-500 dark:text-slate-400 leading-tight">Track<br>Response Times</span>
                            </div>
                            <div class="flex flex-col items-center text-center group/item">
                                <div class="p-2 lg:p-3 rounded-lg lg:rounded-2xl bg-red-500/10 mb-2 lg:mb-3 group-hover/item:scale-110 transition-transform relative">
                                    <i data-lucide="globe" class="w-5 h-5 lg:w-8 lg:h-8 text-blue-500 dark:text-blue-300"></i>
                                    <span class="absolute top-1 lg:top-2 right-1 lg:right-2 flex h-2 lg:h-3 w-2 lg:w-3">
                                        <span class="animate-ping absolute inline-flex h-full w-full rounded-full bg-red-400 opacity-75"></span>
                                        <span class="relative inline-flex rounded-full h-2 lg:h-3 w-2 lg:w-3 bg-red-500"></span>
                                    </span>
                                </div>
                                <span class="text-[9px] lg:text-[11px] font-semibold text-slate-500 dark:text-slate-400 leading-tight">Error Rates</span>
                            </div>
                        </div>

                        <a href="performance_daily.html" 
                           onclick="event.stopPropagation();"
                           class="flex items-center justify-center gap-2 w-full bg-orange-500/10 dark:bg-orange-400/10 text-orange-600 dark:text-orange-400 font-bold py-3 lg:py-4 rounded-lg lg:rounded-2xl text-[10px] lg:text-xs tracking-widest hover:bg-orange-500/20 dark:hover:bg-orange-400/20 border border-orange-500/20 transition-all uppercase mt-auto relative z-10">
                            View Performance Report
                            <i data-lucide="external-link" class="w-4 h-4"></i>
                        </a>
                    </section>
                </div>
            </div>

            <div class="space-y-2 lg:space-y-8">
                <div class="flex items-center gap-4 lg:gap-6">
                    <div class="h-px flex-1 bg-slate-300 dark:bg-slate-700"></div>
                    <h2 class="text-[10px] lg:text-xs font-black uppercase tracking-[0.1em] lg:tracking-[0.4em] text-slate-900 dark:text-slate-100 text-center">SPRINT Execution Results</h2>
                    <div class="h-px flex-1 bg-slate-300 dark:bg-slate-700"></div>
                </div>

                <div class="flex justify-center">
                    <section onclick="window.location.href='performance.html'" 
                             class="sprint-card glass-card card-transition rounded-xl lg:rounded-3xl p-3 lg:p-8 shadow-sm hover:shadow-2xl hover:-translate-y-2 cursor-pointer w-full lg:max-w-lg relative overflow-hidden group">
                        <!-- Card Inner Glow -->
                        <div class="absolute inset-0 bg-gradient-to-br from-white/20 to-transparent opacity-0 group-hover:opacity-100 transition-opacity pointer-events-none"></div>

                        <div class="flex justify-center mb-1 lg:mb-8 relative z-10">
                            <h3 class="text-[11px] lg:text-xs font-bold tracking-widest text-slate-500 dark:text-slate-400 uppercase">Performance Report - SPRINT</h3>
                        </div>

                        <div class="grid grid-cols-3 gap-2 lg:gap-4 mb-2 lg:mb-10 relative z-10">
                            <div class="flex flex-col items-center text-center group/item">
                                <div class="p-2 lg:p-3 rounded-lg lg:rounded-2xl bg-indigo-500/10 mb-2 lg:mb-3 group-hover/item:scale-110 transition-transform">
                                    <i data-lucide="layers" class="w-5 h-5 lg:w-8 lg:h-8 text-indigo-600 dark:text-indigo-400"></i>
                                </div>
                                <span class="text-[9px] lg:text-[11px] font-semibold text-slate-500 dark:text-slate-400 leading-tight">Sprint<br>Analytics</span>
                            </div>
                            <div class="flex flex-col items-center text-center group/item">
                                <div class="p-2 lg:p-3 rounded-lg lg:rounded-2xl bg-purple-500/10 mb-2 lg:mb-3 group-hover/item:scale-110 transition-transform">
                                    <i data-lucide="zap" class="w-5 h-5 lg:w-8 lg:h-8 text-purple-600 dark:text-purple-400"></i>
                                </div>
                                <span class="text-[9px] lg:text-[11px] font-semibold text-slate-500 dark:text-slate-400 leading-tight">Threshold<br>Analysis</span>
                            </div>
                            <div class="flex flex-col items-center text-center group/item">
                                <div class="p-2 lg:p-3 rounded-lg lg:rounded-2xl bg-fuchsia-500/10 mb-2 lg:mb-3 group-hover/item:scale-110 transition-transform">
                                    <i data-lucide="bar-chart-3" class="w-5 h-5 lg:w-8 lg:h-8 text-fuchsia-600 dark:text-fuchsia-400"></i>
                                </div>
                                <span class="text-[9px] lg:text-[11px] font-semibold text-slate-500 dark:text-slate-400 leading-tight">Benchmarking<br>Insight</span>
                            </div>
                        </div>

                        <a href="performance.html" 
                           onclick="event.stopPropagation();"
                           class="flex items-center justify-center gap-2 w-full bg-indigo-500/10 dark:bg-indigo-400/10 text-indigo-600 dark:text-indigo-400 font-bold py-3 lg:py-4 rounded-lg lg:rounded-2xl text-[10px] lg:text-xs tracking-widest hover:bg-indigo-500/20 dark:hover:bg-indigo-400/20 border border-indigo-500/20 transition-all uppercase relative z-10">
                            View Sprint Performance
                            <i data-lucide="external-link" class="w-4 h-4"></i>
                        </a>
                    </section>
                </div>
            </div>

        </main>

        <footer class="mt-auto py-6 text-slate-400 dark:text-slate-600 text-[10px] lg:text-[10px] uppercase tracking-widest text-center w-full">
            &copy; 2026 HirePro Technologies Pvt. Ltd.
        </footer>
    </div>

    <script>
        // Initialize Lucide Icons
        lucide.createIcons();

        // Theme Toggle Functionality with Persistence
        function toggleTheme() {
            const html = document.documentElement;
            const icon = document.getElementById('theme-icon');
            const isDark = html.classList.contains('dark');

            if (isDark) {
                html.classList.remove('dark');
                html.classList.add('light');
                localStorage.setItem('theme', 'light');
                icon.setAttribute('data-lucide', 'moon');
                icon.classList.replace('text-yellow-500', 'text-slate-600');
            } else {
                html.classList.remove('light');
                html.classList.add('dark');
                localStorage.setItem('theme', 'dark');
                icon.setAttribute('data-lucide', 'sun');
                icon.classList.replace('text-slate-600', 'text-yellow-500');
            }
            lucide.createIcons();
        }

        // Sync UI state with detected theme on load
        (function initThemeUI() {
            const html = document.documentElement;
            const icon = document.getElementById('theme-icon');
            if (html.classList.contains('dark')) {
                icon.setAttribute('data-lucide', 'sun');
                icon.classList.add('text-yellow-500');
            } else {
                icon.setAttribute('data-lucide', 'moon');
                icon.classList.add('text-slate-600');
            }
            lucide.createIcons();
        })();
    </script>
</body>
</html>"""
    (Path(__file__).parent.parent / "dashboard.html").write_text(html_content, encoding='utf-8')
    print(f"Landing page generated at {Path(__file__).parent.parent / 'dashboard.html'}")


def generate():
    if not REPORTS_DIR.exists(): return

    # Identify Date Folders
    date_folders = get_report_folders(REPORTS_DIR, limit=5)
    if not date_folders:
        print("No date folders found in reports directory.")
        return

    # Load Expected Reports
    expected_files = []
    if MASTER_LIST_FILE.exists():
        with MASTER_LIST_FILE.open('r') as f: expected_files = [line.strip() for line in f if line.strip()]

    history_data = []  # List of {date, total, passed, failed, reports, missing, no_result, groups}

    # Process Folders (Latest first)
    for i, date_str in enumerate(date_folders):
        folder_path = REPORTS_DIR / date_str
        folder_reports = scan_folder(folder_path, date_str)

        found_filenames = {r['name'] for r in folder_reports}
        missing = [name for name in expected_files if name not in found_filenames]
        new_files = [name for name in found_filenames if name not in expected_files]

        groups = {cat: [] for cat in
                  ["AI - Playwright Reports", "AI - Claude Reports", "API Reports", "CRPO Reports", "ATS Reports",
                   "SSO Reports", "SLOTS Reports", "Microsite Reports", "Cypress Reports"]}
        no_result = []
        for r in folder_reports:
            if r["summary"].get("requests", 0) == 0:
                no_result.append(r)
            else:
                groups[classify_report(r['name'])].append(r)

        # Aggregate metrics (actual individual test case results)
        total = sum(r['summary'].get('requests', 0) for r in folder_reports)
        failed = sum(r['summary'].get('failed', 0) for r in folder_reports)
        passed = total - failed
        pp = round((passed / total * 100), 1) if total > 0 else 0

        history_data.append({
            "date": datetime.datetime.strptime(date_str, '%Y%m%d').strftime('%b %d, %Y'),
            "raw_date": date_str,
            "total": total, "passed": passed, "failed": failed, "pass_percent": pp,
            "reports": folder_reports, "missing": missing, "new_files": new_files, "no_result": no_result,
            "groups": groups
        })

    current_day = history_data[0]
    current_time = datetime.datetime.now().strftime('%Y-%m-%d %H:%M:%S');
    commit_id = get_git_commit()

    # Generate History UI Components
    history_table_rows = []
    date_options = []

    for h in history_data:
        health_cls = "health-good" if h["pass_percent"] > 95 else (
            "health-fair" if h["pass_percent"] > 85 else "health-bad")
        history_table_rows.append(f"""
            <tr id="row-{h['raw_date']}">
                <td class="history-date">{h['date']}</td>
                <td>{h['total']} / <span style="color: var(--text-muted); font-size: 0.8rem; opacity: 0.6;">{TARGET_EXECUTION_GOAL}</span></td>
                <td style="color: var(--success); font-weight: 700;">{h['passed']}</td>
                <td style="color: var(--danger); font-weight: 700;">{h['failed']}</td>
                <td><span class="health-index {health_cls}">{h['pass_percent']}%</span></td>
                <td><button class="view-btn {'active' if h == current_day else ''}" onclick="switchDate('{h['raw_date']}', true)">View Details</button></td>
            </tr>""")
        date_options.append(f'<option value="{h["raw_date"]}">{h["date"]}</option>')

    # Generate Day-Specific Breakdowns
    day_views_html = []
    today_str = datetime.datetime.now().strftime('%Y%m%d')
    for idx, h in enumerate(history_data):
        is_active = "active" if h == current_day else ""
        cov = round((h["total"] / TARGET_EXECUTION_GOAL) * 100, 1) if TARGET_EXECUTION_GOAL > 0 else 0

        # Only show the red stale alert for the VERY LATEST report if it's not today
        # We don't want history reports to show red just because they are in the past
        is_stale = (idx == 0 and h['raw_date'] != today_str)
        stale_cls = "hero-stale" if is_stale else ""

        view_parts = [f"""<div id="view-{h['raw_date']}" class="day-view {is_active}">"""]

        # Hero Banner
        view_parts.append(f"""
        <div class="hero-banner {'stale-banner' if is_stale else ''}">
            <div class="hero-main">
                <div class="hero-label">
                    Execution Goal ({h['date']})
                    {'<span class="sync-badge"><i class="fas fa-clock-rotate-left"></i> OUT OF SYNC</span>' if is_stale else ''}
                </div>
                <div class="hero-value">{h['total']} <span class="hero-target">/ {TARGET_EXECUTION_GOAL} Test Cases</span></div>
                {f'<span class="fail-badge">{len(h["no_result"]) + len(h["missing"])} Failed Test (s)</span>' if (h["no_result"] or h["missing"]) else ""}
            </div>
            <div class="progress-container"><div class="progress-header"><span>Lifecycle Coverage</span><span>{cov}% Complete</span></div>
            <div class="progress-outer"><div class="progress-inner" style="width: {cov}%"></div></div></div>
        </div>
        <div class="stats-grid">
            <div class="stat-card blue"><div class="stat-icon icon-blue"><i class="fas fa-microscope"></i></div><div class="stat-info"><span class="value">{h['total']}</span><span class="label">Executed</span></div></div>
            <div class="stat-card green"><div class="stat-icon icon-green"><i class="fas fa-check-circle"></i></div><div class="stat-info"><span class="value">{h['passed']}</span><span class="label">Passed</span></div></div>
            <div class="stat-card red"><div class="stat-icon icon-red"><i class="fas fa-times-circle"></i></div><div class="stat-info"><span class="value">{h['failed']}</span><span class="label">Failed</span></div></div>
            <div class="stat-card orange">
                <div class="stat-icon icon-orange"><i class="fas fa-folder-open"></i></div>
                <div class="stat-info">
                    <span class="value">{len(h['reports'])} / {EXPECTED_REPORT_COUNT}</span>
                    <span class="label">TEST REPORTS</span>
                    {f'<div class="sync-label error"><i class="fas fa-exclamation-circle"></i> OUT OF SYNC</div>' if len(h['reports']) < EXPECTED_REPORT_COUNT else (f'<div class="sync-label error"><i class="fas fa-exclamation-circle"></i> NEW REPORTS DETECTED</div>' if len(h['reports']) > EXPECTED_REPORT_COUNT else '<div class="sync-label" style="color:#059669"><i class="fas fa-check-double"></i> ALL SYNCED</div>')}
                </div>
            </div>
        </div>

        <div class="search-area">
            <i class="fas fa-search"></i>
            <input type="text" placeholder="Search tests for {h['date']}..." onkeyup="searchInside(this)">
        </div>

        """)
        if h["no_result"]:
            view_parts.append(
                f"""<div class="alert-row"><div class="alert-card"><div style="display:flex; align-items:center;"><i class="fas fa-exclamation-triangle"></i><span style="font-weight:700;">{len(h['no_result'])} Critical: Report(s) found with Zero results (Possible execution stall or crash) | {h['date']}</span></div><span class="badge-urgent">Action Required</span></div></div>""")

        # Zero Result Reports
        if h["no_result"]:
            h["no_result"].sort(key=lambda x: x['mod_time_ts'], reverse=True)
            view_parts.append(
                f"""<details class="report-group" style="border-color: #feb2b2; margin-bottom: 2.5rem;"><summary style="background: #fff5f5; color: #c53030; display:flex; align-items:center;"><div class="collapsible-loader"></div><span><i class="fas fa-bug"></i>Critical: Incomplete Executions ({len(h['no_result'])})</span></summary><div class="content-wrapper"><div class="content-inner"><table><thead><tr><th>Identity</th><th>Format</th><th>Check Time</th><th>Status</th></tr></thead><tbody>""")
            for r in h["no_result"]: view_parts.append(
                f"""<tr class="report-row"><td><a href="{r.get('view_path', r['path'])}" class="report-link" target="_blank" style="color:#c53030;"><i class="fas fa-exclamation-circle"></i> <span>{r['name']}</span></a></td><td><span class="badge badge-html">{r['type']}</span></td><td style="font-size: 0.85rem; color: var(--text-muted); font-weight: 500;">{r['mod_time']}</td><td style="color: #c53030; font-style: italic; font-weight: 700;">No metrics found</td></tr>""")
            view_parts.append("</tbody></table></div></div></details>")

        # Missing Reports
        if h["missing"]:
            view_parts.append(
                f"""<details class="report-group" style="border-color: #f59e0b; margin-bottom: 2.5rem;"><summary style="background: #fffbeb; color: #d97706; display:flex; align-items:center;"><div class="collapsible-loader"></div><span><i class="fas fa-file-circle-exclamation"></i> Identity Integrity Check: Missing ({len(h['missing'])})</span></summary><div class="content-wrapper"><div class="content-inner"><table><thead><tr><th>Expected Filename</th><th>Status</th></tr></thead><tbody>""")
            for name in h["missing"]: view_parts.append(
                f"""<tr class="report-row"><td><span class="report-link" style="color:#d97706;"><i class="fas fa-file-circle-xmark"></i> <span>{name}</span></span></td><td style="color: #d97706; font-style: italic; font-weight: 700;">Missing from manifest</td></tr>""")
            view_parts.append("</tbody></table></div></div></details>")

        # New Reports (Not in Master List)
        if h.get("new_files"):
            view_parts.append(
                f"""<details class="report-group" style="border-color: #10b981; margin-bottom: 2.5rem;"><summary style="background: #ecfdf5; color: #059669; display:flex; align-items:center;"><div class="collapsible-loader"></div><span><i class="fas fa-file-circle-plus"></i> Integrity Insight: Newly Detected Reports ({len(h['new_files'])})</span></summary><div class="content-wrapper"><div class="content-inner"><table><thead><tr><th>Detected Filename</th><th>Status</th></tr></thead><tbody>""")
            for name in h["new_files"]:
                view_parts.append(
                    f"""<tr class="report-row"><td><span class="report-link" style="color:#059669;"><i class="fas fa-plus-circle"></i> <span>{name}</span></span></td><td style="color: #059669; font-style: italic; font-weight: 700;">New report added</td></tr>""")
            view_parts.append("</tbody></table></div></div></details>")

        # Normal Groups
        for group, items in h["groups"].items():
            if not items: continue
            items.sort(key=lambda x: x['mod_time_ts'], reverse=True)
            g_total = sum(i['summary'].get('requests', 0) for i in items);
            g_failed = sum(i['summary'].get('failed', 0) for i in items);
            g_passed = g_total - g_failed
            view_parts.append(
                f"""<details class="report-group"><summary style="background: #f8fafc; display:flex; align-items:center;"><div class="collapsible-loader"></div><div style="display:flex; align-items:center; flex:1; width:100%;"><span style="font-weight:800; color:var(--text);">{group} ({len(items)})</span><div style="display:flex; gap:1.25rem; font-size:0.75rem; margin-left:auto; margin-right: 1.5rem;"><span class="chip" data-tooltip="TOTAL">{g_total} T</span><span class="chip success" data-tooltip="PASSED">{g_passed} P</span><span class="chip danger" data-tooltip="FAILED" style="{'background:#ef4444; color:white;' if g_failed > 0 else ''}">{g_failed} F</span></div></div></summary><div class="content-wrapper"><div class="content-inner"><table><thead><tr><th>Suite Identity</th><th>Format</th><th>Verification</th><th>Quality Insight</th></tr></thead><tbody>""")
            for r in items:
                s = r['summary'];
                pas, fld, pct = s['requests'] - s['failed'], s['failed'], s['pass_percent'];
                clr = "#10b981" if pct > 90 else ("#f59e0b" if pct > 70 else "#ef4444")
                failed_list = ""
                if s.get("failed_cases"):
                    cases_str = ", ".join(s["failed_cases"][:5])
                    if len(s["failed_cases"]) > 5: cases_str += "..."
                    failed_list = f'<div style="font-size:0.65rem; color:#ef4444; margin-top:4px; font-weight:600;"><i class="fas fa-bug"></i> {cases_str}</div>'

                view_parts.append(
                    f"""<tr class="report-row"><td><a href="{r.get('view_path', r['path'])}" class="report-link" target="_blank"><i class="fas {'fa-file-lines' if r['type'] == 'HTML' else 'fa-file-excel'}" style="color:var(--primary)"></i> <span>{r['name']}</span></a></td><td><span class="badge {'badge-html' if r['type'] == 'HTML' else 'badge-excel'}">{r['type']}</span></td><td style="font-size: 0.85rem; color: var(--text-muted); font-weight: 500;">{r['mod_time']}</td><td><div style="display:flex;gap:4px;margin-bottom:6px;"><span class="chip">{s['requests']} T</span><span class="chip success">{pas} P</span><span class="chip danger">{fld} F</span></div><div style="display:flex; align-items:center; gap:8px;"><div style="height:6px;width:100px;background:#f1f5f9;border-radius:3px;overflow:hidden;"><div style="height:100%;width:{pct}%;background:{clr}"></div></div><span style="font-size:0.75rem;font-weight:700;color:{clr}">{pct}%</span></div>{failed_list}</td></tr>""")
            view_parts.append("</tbody></table></div></div></details>")

        view_parts.append("</div>")
        day_views_html.append("".join(view_parts))

    html_parts = [f"""<!DOCTYPE html><html lang="en"><head><meta charset="UTF-8"><meta name="viewport" content="width=device-width, initial-scale=1.0">
    <link rel="icon" type="image/png" href="https://lh3.googleusercontent.com/proxy/98it6d0vGqaPttxE8ImrHhVvaz5XpPeZ7qiXHcnm9PiPrwTBGvZcWp2vdQVdgZt5b7Vfu7kXnkh6mrKs_q_JIE5GGlw8uRAOeSvJOEtgNXWrXgOoGjGFCnKCA1gITLLZKNxv1mV6szDHEnNLK7RbJnfk-eFMZPlGXRpU2iKeBGr2Gm3-i_Lnv-IVisLlJwRR55nNMx9HndfYnmLOlraDHGgp9Rc7V4pNO1N8S1ZugYR5SUVMi8K_WGFwvYWsEh2cEnW6x-Zw-ncSM27yX509d6pmuUvfohenPAxHMNORCeWO">
    <title>HirePro Quality Dashboard</title><link href="https://fonts.googleapis.com/css2?family=Outfit:wght@300;400;500;600;700;800&display=swap" rel="stylesheet">
    <link rel="stylesheet" href="https://cdnjs.cloudflare.com/ajax/libs/font-awesome/6.4.0/css/all.min.css">
    {generate_styles()}</head><body>
    <div class="container">
    <header>
        <div style="display: flex; align-items: center; gap: 1rem;">
            <a href="dashboard.html" class="header-home-btn">
                <i class="fas fa-home"></i>
            </a>
            <a href="dashboard.html" style="text-decoration: none;">
                <img src="https://hirepro.in/wp-content/uploads/2025/05/HirePro-logo.svg" alt="HirePro Logo" style="height: 32px; width: auto;">
            </a>
        </div>
        <div style="text-align: center;"><h1>Quality Dashboard</h1><p class="subtitle">Unified Test Lifecycle Monitoring</p></div>
        <div style="text-align: right;">
            <div style="font-size: 0.8rem; color: var(--text-muted); font-weight: 600; margin-bottom: 0.25rem;">LATEST AUDIT</div>
            <div style="font-weight: 700;"><i class="fas fa-sync-alt" style="color:var(--success); margin-right: 0.5rem;"></i> {current_time}</div>
        </div>
    </header>

    <div class="main-nav">
        <div class="nav-indicator"></div>
        <button class="nav-btn active" id="btn-current" onclick="showTab('current')"><i class="fas fa-satellite-dish"></i> Live Audit</button>
        <button class="nav-btn" id="btn-history" onclick="showTab('history')"><i class="fas fa-clock-rotate-left"></i> Quality History</button>
    </div>

    <div id="dashboard-main" style="position: relative;">
        <div id="loader" class="loader-overlay">
            <div style="text-align: center;">
                <div class="loader-spinner" style="margin: 0 auto;"></div>
                <div class="loader-text">Syncing Dashboard...</div>
            </div>
        </div>
        <div id="tab-history" class="tab-content">
            <div class="history-section">
                <div class="trend-title" style="margin-bottom: 1.5rem;"><i class="fas fa-history"></i> Historical Audit Timeline</div>
                <div class="history-table-wrapper">
                    <table class="history-table">
                        <thead><tr><th>Execution Date</th><th>Total Cases</th><th>Passed</th><th>Failed</th><th>Success Rate</th><th>Action</th></tr></thead>
                        <tbody>{"".join(history_table_rows)}</tbody>
                    </table>
                </div>
            </div>
            <div class="date-selector-wrapper">
                <span style="font-weight: 800; color: var(--text-muted); font-size: 0.8rem; text-transform: uppercase;">Switch Analytical Context</span>
                <select id="dateSelector" class="date-select-input" onchange="switchDate(this.value)">
                    {"".join(date_options)}
                </select>
            </div>
        </div>

        <div id="tab-current" class="tab-content active">
            <div id="dayViewsContainer">
                {"".join(day_views_html)}
            </div>
        </div>
    </div>
    """]

    html_parts.append(f"""
    <a href="dashboard.html" class="floating-home" data-tooltip="Return to Menu">
        <i class="fas fa-home"></i>
    </a>
    <button class="floating-reset" onclick="switchDate('{current_day['raw_date']}', true)" data-tooltip="Reset to Live Audit">
        <i class="fas fa-rotate-left"></i>
    </button>
    <script>
        function showTab(tabId) {{
            const indicator = document.querySelector('.nav-indicator');
            const targetBtn = document.getElementById('btn-' + tabId);

            if (targetBtn && indicator) {{
                indicator.style.width = targetBtn.offsetWidth + 'px';
                indicator.style.left = targetBtn.offsetLeft + 'px';
            }}

            document.querySelectorAll('.tab-content').forEach(t => t.classList.remove('active'));
            document.querySelectorAll('.nav-btn').forEach(b => b.classList.remove('active'));
            document.getElementById('tab-' + tabId).classList.add('active');
            if(targetBtn) targetBtn.classList.add('active');
        }}
        // Initialize Nav
        function updateIndicator() {{
            const activeBtn = document.querySelector('.nav-btn.active');
            const indicator = document.querySelector('.nav-indicator');
            if (activeBtn && indicator) {{
                indicator.style.width = activeBtn.offsetWidth + 'px';
                indicator.style.left = activeBtn.offsetLeft + 'px';
            }}
        }}
        window.addEventListener('load', () => {{
            updateIndicator();
            const loader = document.getElementById('loader');
            setTimeout(() => {{
                loader.style.opacity = '0';
                setTimeout(() => {{ loader.style.display = 'none'; loader.style.opacity = '1'; }}, 300);
            }}, 500);
            setTimeout(updateIndicator, 200); // Safety double-check after layout settles
        }});
        window.addEventListener('resize', updateIndicator);
        function switchDate(dateStr, toTop) {{
            document.querySelectorAll('.day-view').forEach(v => v.classList.remove('active'));
            document.querySelectorAll('.view-btn').forEach(b => b.classList.remove('active'));

            const targetView = document.getElementById('view-' + dateStr);
            if (targetView) {{
                targetView.classList.add('active');
                // Auto-collapse any open sections when switching
                targetView.querySelectorAll('details[open]').forEach(d => {{
                    d.open = false;
                    const w = d.querySelector('.content-wrapper');
                    if(w) w.style.gridTemplateRows = '';
                }});
            }}

            let btn = document.querySelector('#row-' + dateStr + ' .view-btn');
            if(btn) btn.classList.add('active');
            document.getElementById('dateSelector').value = dateStr;
            showTab('current');
            window.scrollTo({{ top: toTop ? 0 : 350, behavior: 'smooth' }});
        }}
        function searchInside(input) {{
            let filter = input.value.toUpperCase();
            let parent = input.closest('.day-view');
            parent.querySelectorAll('.report-group').forEach(group => {{
                let rows = group.querySelectorAll('.report-row'); let foundAny = false;
                rows.forEach(row => {{ let text = row.innerText.toUpperCase(); if (text.indexOf(filter) > -1) {{ row.style.display = ""; foundAny = true; }} else {{ row.style.display = "none"; }} }});
                group.style.display = foundAny ? "" : "none";
            }});
        }}
        document.querySelectorAll('details').forEach(details => {{
            const summary = details.querySelector('summary'), wrapper = details.querySelector('.content-wrapper');
            summary.addEventListener('click', e => {{
                e.preventDefault();
                if (summary.classList.contains('loading')) return;

                const isOpening = !details.open;

                if (!isOpening) {{
                    summary.classList.add('loading');
                    wrapper.style.gridTemplateRows = '0fr';
                    const onFinish = () => {{
                        details.open = false;
                        wrapper.style.gridTemplateRows = '';
                        summary.classList.remove('loading');
                        wrapper.removeEventListener('transitionend', onFinish);
                    }};
                    wrapper.addEventListener('transitionend', onFinish, {{ once: true }});
                    setTimeout(onFinish, 500); // Safety fallback
                }} else {{
                    details.open = true;
                }}
            }});
        }});
        document.querySelectorAll('a.report-link:not([target="_blank"])').forEach(link => {{
            link.addEventListener('click', () => {{
                const loader = document.getElementById('loader');
                loader.style.display = 'flex';
                setTimeout(() => {{ loader.style.display = 'none'; }}, 1000);
            }});
        }});
    </script><div class="container"><footer><div style="display: flex; align-items: center; gap: 0.75rem;"><span style="text-transform: uppercase; font-size: 0.7rem; letter-spacing: 0.05em; font-weight: 700;">Build ID:</span><span class="commit-badge"><i class="fas fa-code-branch" style="margin-right:0.4rem; opacity:0.5;"></i>{commit_id}</span></div><div>&copy; 2026 HirePro . All rights reserved.</div></footer></div></body></html>""")
    OUTPUT_FILE.write_text("".join(html_parts), encoding='utf-8');
    generate_landing_page()
    print(f"Dashboard generated at {OUTPUT_FILE}")


if __name__ == "__main__": generate()